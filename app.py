import openpyxl as xl
from openpyxl.chart import LineChart, BarChart, Reference
from pathlib import Path


from datetime import date

p = Path()

for file in p.glob('*.xlsx'):

        wb = xl.load_workbook(file)
        ws = wb.copy_worksheet(wb.active)
        ws.title = 'Processed on ' + str(date.today())
        ws.cell(1,ws.max_column + 1).value = 'Discounted Price'
        for row in range(2, ws.max_row + 1):
            price = ws.cell(row,3)
            discounted_price = price.value * 0.9
            ws.cell(row, 4).value = discounted_price
        print(file," is processed successfully. Check the workbook for more details.")

ref_1 = Reference(ws,min_col=4, min_row=2, max_col=4, max_row=51)

ls = LineChart()
ls.add_data(ref_1)
ws.add_chart(ls,'g2')


bc = BarChart()
bc.add_data(ref_1)
ws.add_chart(bc,'g20')


wb.save(file)












